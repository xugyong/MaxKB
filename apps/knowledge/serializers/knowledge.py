import io
import json
import os
import pickle
import re
import tempfile
import traceback
import zipfile
from collections import defaultdict
from functools import reduce
from tempfile import TemporaryDirectory
from typing import Dict, List
from urllib.parse import quote

import requests
import uuid_utils.compat as uuid
from celery_once import AlreadyQueued
from django.core import validators
from django.core.files.uploadedfile import SimpleUploadedFile
from django.db import transaction, models
from django.db.models import QuerySet
from django.db.models.functions import Reverse, Substr
from django.db.models.query_utils import Q
from django.http import HttpResponse
from django.utils.translation import gettext_lazy as _, gettext
from rest_framework import serializers

from common.config.embedding_config import VectorStore
from common.database_model_manage.database_model_manage import DatabaseModelManage
from common.db.search import native_search, get_dynamics_model, native_page_search
from common.db.sql_execute import select_list
from common.event.listener_manage import ListenerManagement
from common.exception.app_exception import AppApiException
from common.field.common import UploadedFileField
from common.utils.common import post, get_file_content, parse_image, bulk_create_in_batches
from common.utils.fork import Fork, ChildLink
from common.utils.logger import maxkb_logger
from common.utils.split_model import get_split_model
from knowledge.models import Knowledge, KnowledgeScope, KnowledgeType, Document, Paragraph, Problem, \
    ProblemParagraphMapping, TaskType, State, SearchMode, KnowledgeFolder, File, Tag, DocumentTag, KnowledgeWorkflow, \
    FileSourceType
from knowledge.serializers.common import BatchSerializer, BatchMoveSerializer, ProblemParagraphObject
from knowledge.serializers.common import ProblemParagraphManage, drop_knowledge_index, \
    get_embedding_model_id_by_knowledge_id, MetaSerializer, \
    GenerateRelatedSerializer, get_embedding_model_by_knowledge_id, list_paragraph, write_image, zip_dir, \
    update_resource_mapping_by_knowledge
from knowledge.serializers.document import DocumentSerializers
from knowledge.task.embedding import embedding_by_knowledge, delete_embedding_by_knowledge
from knowledge.task.generate import generate_related_by_knowledge_id
from knowledge.task.sync import sync_web_knowledge, sync_replace_web_knowledge
from maxkb.conf import PROJECT_DIR
from maxkb.const import CONFIG
from models_provider.models import Model
from system_manage.models import WorkspaceUserResourcePermission, AuthTargetType
from system_manage.models.resource_mapping import ResourceMapping
from system_manage.serializers.resource_mapping_serializers import ResourceMappingSerializer
from system_manage.serializers.user_resource_permission import UserResourcePermissionSerializer
from users.serializers.user import is_workspace_manage


class KnowledgeModelSerializer(serializers.ModelSerializer):
    class Meta:
        model = Knowledge
        fields = ['id', 'name', 'desc', 'meta', 'folder_id', 'type', 'workspace_id', 'create_time',
                  'update_time', 'file_size_limit', 'file_count_limit', 'embedding_model_id']


class KnowledgeBaseCreateRequest(serializers.Serializer):
    name = serializers.CharField(required=True, label=_('knowledge name'))
    folder_id = serializers.CharField(required=True, label=_('folder id'))
    desc = serializers.CharField(required=False, allow_null=True, allow_blank=True, label=_('knowledge description'))
    embedding_model_id = serializers.CharField(required=True, label=_('knowledge embedding'))


class KnowledgeImportRequest(serializers.Serializer):
    file = UploadedFileField(required=True, label=_("file"))


class KnowledgeWebCreateRequest(serializers.Serializer):
    name = serializers.CharField(required=True, label=_('knowledge name'))
    folder_id = serializers.CharField(required=True, label=_('folder id'))
    desc = serializers.CharField(required=False, allow_null=True, allow_blank=True, label=_('knowledge description'))
    embedding_model_id = serializers.CharField(required=True, label=_('knowledge embedding'))
    source_url = serializers.CharField(required=True, label=_('source url'))
    selector = serializers.CharField(required=False, label=_('knowledge selector'), allow_null=True, allow_blank=True)


class KnowledgeEditRequest(serializers.Serializer):
    name = serializers.CharField(required=False, max_length=64, min_length=1, label=_('knowledge name'))
    desc = serializers.CharField(required=False, max_length=256, min_length=1, label=_('knowledge description'))
    meta = serializers.DictField(required=False)
    application_id_list = serializers.ListSerializer(
        required=False,
        child=serializers.UUIDField(required=True, label=_('application id')),
        label=_('application id list')
    )
    file_size_limit = serializers.IntegerField(required=False, label=_('file size limit'))
    file_count_limit = serializers.IntegerField(required=False, label=_('file count limit'))

    @staticmethod
    def get_knowledge_meta_valid_map():
        knowledge_meta_valid_map = {
            KnowledgeType.BASE: MetaSerializer.BaseMeta,
            KnowledgeType.WEB: MetaSerializer.WebMeta
        }
        return knowledge_meta_valid_map

    def is_valid(self, *, knowledge: Knowledge = None):
        super().is_valid(raise_exception=True)
        if 'meta' in self.data and self.data.get('meta') is not None:
            knowledge_meta_valid_map = self.get_knowledge_meta_valid_map()
            valid_class = knowledge_meta_valid_map.get(knowledge.type)
            valid_class(data=self.data.get('meta')).is_valid(raise_exception=True)


class HitTestSerializer(serializers.Serializer):
    query_text = serializers.CharField(required=True, label=_('query text'))
    top_number = serializers.IntegerField(required=True, max_value=10000, min_value=1, label=_("top number"))
    similarity = serializers.FloatField(required=True, max_value=2, min_value=0, label=_('similarity'))
    search_mode = serializers.CharField(required=True, label=_('search mode'), validators=[
        validators.RegexValidator(regex=re.compile("^embedding|keywords|blend$"),
                                  message=_('The type only supports embedding|keywords|blend'), code=500)
    ])


class KnowledgeSerializer(serializers.Serializer):
    class Query(serializers.Serializer):
        workspace_id = serializers.CharField(required=True)
        folder_id = serializers.CharField(required=False, label=_('folder id'), allow_null=True)
        name = serializers.CharField(required=False, label=_('knowledge name'), allow_null=True, allow_blank=True,
                                     max_length=64, min_length=1)
        desc = serializers.CharField(required=False, label=_('knowledge description'), allow_null=True,
                                     allow_blank=True, max_length=256, min_length=1)
        user_id = serializers.UUIDField(required=False, label=_('user id'), allow_null=True)
        scope = serializers.CharField(required=False, label=_('knowledge scope'), allow_null=True)
        create_user = serializers.UUIDField(required=False, label=_('create user'), allow_null=True)

        @staticmethod
        def is_x_pack_ee():
            workspace_user_role_mapping_model = DatabaseModelManage.get_model("workspace_user_role_mapping")
            role_permission_mapping_model = DatabaseModelManage.get_model("role_permission_mapping_model")
            return workspace_user_role_mapping_model is not None and role_permission_mapping_model is not None

        def get_query_set(self, workspace_manage, is_x_pack_ee):
            self.is_valid(raise_exception=True)
            workspace_id = self.data.get("workspace_id")
            query_set_dict = {}
            query_set = QuerySet(model=get_dynamics_model({
                'temp.name': models.CharField(),
                'temp.desc': models.CharField(),
                "document_temp.char_length": models.IntegerField(),
                'temp.create_time': models.DateTimeField(),
                'temp.user_id': models.CharField(),
                'temp.workspace_id': models.CharField(),
                'temp.folder_id': models.CharField(),
                'temp.id': models.CharField(),
                'temp.scope': models.CharField(),
            }))
            folder_query_set = QuerySet(KnowledgeFolder)

            if "desc" in self.data and self.data.get('desc') is not None:
                query_set = query_set.filter(**{'temp.desc__icontains': self.data.get("desc")})
                folder_query_set = folder_query_set.filter(**{'desc__icontains': self.data.get("desc")})
            if "name" in self.data and self.data.get('name') is not None:
                query_set = query_set.filter(**{'temp.name__icontains': self.data.get("name")})
                folder_query_set = folder_query_set.filter(**{'name__icontains': self.data.get("name")})
            if "workspace_id" in self.data and self.data.get('workspace_id') is not None:
                query_set = query_set.filter(**{'temp.workspace_id': self.data.get("workspace_id")})
                folder_query_set = folder_query_set.filter(**{'workspace_id': self.data.get("workspace_id")})
            if "folder_id" in self.data and self.data.get('folder_id') is not None and self.data.get(
                    'workspace_id') != self.data.get('folder_id'):
                query_set = query_set.filter(**{'temp.folder_id': self.data.get("folder_id")})
                folder_query_set = folder_query_set.filter(**{'parent_id': self.data.get("folder_id")})
            if "scope" in self.data and self.data.get('scope') is not None:
                query_set = query_set.filter(**{'temp.scope': self.data.get("scope")})
            if "create_user" in self.data and self.data.get('create_user') is not None:
                query_set = query_set.filter(**{'temp.user_id': self.data.get("create_user")})
            query_set = query_set.order_by("-temp.create_time", "temp.id")
            query_set_dict['default_sql'] = query_set

            query_set_dict['knowledge_custom_sql'] = QuerySet(model=get_dynamics_model({
                'knowledge.workspace_id': models.CharField(),
            })).filter(**{'knowledge.workspace_id': workspace_id})
            # query_set_dict['folder_query_set'] = folder_query_set
            if not workspace_manage:
                query_set_dict['workspace_user_resource_permission_query_set'] = QuerySet(
                    WorkspaceUserResourcePermission).filter(
                    auth_target_type="KNOWLEDGE",
                    workspace_id=workspace_id,
                    user_id=self.data.get("user_id"))
            return query_set_dict

        def page(self, current_page: int, page_size: int):
            self.is_valid(raise_exception=True)
            folder_id = self.data.get('folder_id', self.data.get("workspace_id"))
            root = KnowledgeFolder.objects.filter(id=folder_id).first()
            if not root:
                raise serializers.ValidationError(_('Folder not found'))
            workspace_manage = is_workspace_manage(self.data.get('user_id'), self.data.get('workspace_id'))
            is_x_pack_ee = self.is_x_pack_ee()
            result = native_page_search(
                current_page,
                page_size,
                self.get_query_set(workspace_manage, is_x_pack_ee),
                select_string=get_file_content(
                    os.path.join(
                        PROJECT_DIR,
                        "apps",
                        "knowledge", 'sql',
                        'list_knowledge.sql' if workspace_manage else (
                            'list_knowledge_user_ee.sql' if is_x_pack_ee else 'list_knowledge_user.sql'
                        )
                    )
                ),
                post_records_handler=lambda r: r
            )
            return ResourceMappingSerializer().get_resource_count(result)

        def list(self):
            self.is_valid(raise_exception=True)
            folder_id = self.data.get('folder_id')
            if not folder_id:
                folder_id = self.data.get('workspace_id')
            root = KnowledgeFolder.objects.filter(id=folder_id).first()
            if not root:
                raise serializers.ValidationError(_('Folder not found'))
            workspace_manage = is_workspace_manage(self.data.get('user_id'), self.data.get('workspace_id'))
            is_x_pack_ee = self.is_x_pack_ee()
            return native_search(
                self.get_query_set(workspace_manage, is_x_pack_ee),
                select_string=get_file_content(
                    os.path.join(
                        PROJECT_DIR,
                        "apps",
                        "knowledge", 'sql',
                        'list_knowledge.sql' if workspace_manage else (
                            'list_knowledge_user_ee.sql' if self.is_x_pack_ee() else 'list_knowledge_user.sql'
                        )
                    )
                ),
            )

    class Operate(serializers.Serializer):
        user_id = serializers.UUIDField(required=True, label=_('user id'))
        workspace_id = serializers.CharField(required=True, label=_('workspace id'))
        knowledge_id = serializers.UUIDField(required=True, label=_('knowledge id'))

        def is_valid(self, *, raise_exception=False):
            super().is_valid(raise_exception=True)
            workspace_id = self.data.get('workspace_id')
            query_set = QuerySet(Knowledge).filter(id=self.data.get('knowledge_id'))
            if workspace_id:
                query_set = query_set.filter(workspace_id=workspace_id)
            if not query_set.exists():
                raise AppApiException(500, _('Knowledge id does not exist'))

        @transaction.atomic
        def embedding(self, with_valid=True):
            if with_valid:
                self.is_valid(raise_exception=True)
            knowledge_id = self.data.get('knowledge_id')
            knowledge = QuerySet(Knowledge).filter(id=knowledge_id).first()
            embedding_model_id = knowledge.embedding_model_id
            embedding_model = QuerySet(Model).filter(id=embedding_model_id).first()
            if embedding_model is None:
                raise AppApiException(500, _('Model does not exist'))
            ListenerManagement.update_status(
                QuerySet(Document).filter(knowledge_id=self.data.get('knowledge_id')),
                TaskType.EMBEDDING,
                State.PENDING
            )
            ListenerManagement.update_status(
                QuerySet(Paragraph).filter(knowledge_id=self.data.get('knowledge_id')),
                TaskType.EMBEDDING,
                State.PENDING
            )
            ListenerManagement.get_aggregation_document_status_by_knowledge_id(self.data.get('knowledge_id'))()
            embedding_model_id = get_embedding_model_id_by_knowledge_id(self.data.get('knowledge_id'))
            try:
                embedding_by_knowledge.delay(knowledge_id, embedding_model_id)
            except AlreadyQueued as e:
                raise AppApiException(500, _('Failed to send the vectorization task, please try again later!'))

        def generate_related(self, instance: Dict, with_valid=True):
            if with_valid:
                self.is_valid(raise_exception=True)
                GenerateRelatedSerializer(data=instance).is_valid(raise_exception=True)
            knowledge_id = self.data.get('knowledge_id')
            model_id = instance.get("model_id")
            prompt = instance.get("prompt")
            model_params_setting = instance.get("model_params_setting")
            state_list = instance.get('state_list')
            ListenerManagement.update_status(
                QuerySet(Document).filter(knowledge_id=knowledge_id),
                TaskType.GENERATE_PROBLEM,
                State.PENDING
            )
            ListenerManagement.update_status(
                QuerySet(Paragraph).annotate(
                    reversed_status=Reverse('status'),
                    task_type_status=Substr('reversed_status', TaskType.GENERATE_PROBLEM.value, 1),
                ).filter(
                    task_type_status__in=state_list, knowledge_id=knowledge_id
                ).values('id'),
                TaskType.GENERATE_PROBLEM,
                State.PENDING
            )
            ListenerManagement.get_aggregation_document_status_by_knowledge_id(knowledge_id)()
            try:
                generate_related_by_knowledge_id.delay(knowledge_id, model_id, model_params_setting, prompt, state_list)
            except AlreadyQueued as e:
                raise AppApiException(500, _('Failed to send the vectorization task, please try again later!'))

        def list_application(self, with_valid=True):
            if with_valid:
                self.is_valid(raise_exception=True)
            # knowledge = QuerySet(Knowledge).get(id=self.data.get("knowledge_id"))
            return select_list(
                get_file_content(
                    os.path.join(PROJECT_DIR, "apps", "knowledge", 'sql', 'list_knowledge_application.sql')
                ),
                [
                    self.data.get('user_id'),
                ]
            )

        @staticmethod
        def is_x_pack_ee():
            workspace_user_role_mapping_model = DatabaseModelManage.get_model("workspace_user_role_mapping")
            role_permission_mapping_model = DatabaseModelManage.get_model("role_permission_mapping_model")
            return workspace_user_role_mapping_model is not None and role_permission_mapping_model is not None

        def one(self):
            self.is_valid()
            workspace_manage = is_workspace_manage(self.data.get('user_id'), self.data.get('workspace_id'))
            is_x_pack_ee = self.is_x_pack_ee()

            query_set_dict = {
                'default_sql': QuerySet(
                    model=get_dynamics_model({'temp.id': models.CharField()})
                ).filter(**{'temp.id': self.data.get("knowledge_id")}),
                'knowledge_custom_sql': QuerySet(
                    model=get_dynamics_model({'knowledge.id': models.CharField()})
                ).filter(**{'knowledge.id': self.data.get("knowledge_id")}),
            }
            if not workspace_manage:
                query_set_dict['workspace_user_resource_permission_query_set'] = QuerySet(
                    WorkspaceUserResourcePermission).filter(
                    auth_target_type="KNOWLEDGE",
                    workspace_id=self.data.get('workspace_id'),
                    user_id=self.data.get("user_id")
                )
            all_application_list = [str(adm.get('id')) for adm in self.list_application(with_valid=False)]
            knowledge_dict = native_search(query_set_dict, select_string=get_file_content(
                os.path.join(
                    PROJECT_DIR, "apps", "knowledge", 'sql',
                    'list_knowledge.sql' if workspace_manage else (
                        'list_knowledge_user_ee.sql' if is_x_pack_ee else 'list_knowledge_user.sql'
                    )
                )
            ), with_search_one=True)
            workflow = {}

            if knowledge_dict.get('type') == 4:
                from knowledge.models import KnowledgeWorkflow
                k = QuerySet(KnowledgeWorkflow).filter(knowledge_id=knowledge_dict.get('id')).first()
                if k:
                    workflow['work_flow'] = k.work_flow
                    workflow['is_publish'] = k.is_publish
                    workflow['publish_time'] = k.publish_time
            return {
                **knowledge_dict,
                **workflow,
                'meta': json.loads(knowledge_dict.get('meta', '{}')),
                'application_id_list': list(filter(
                    lambda application_id: all_application_list.__contains__(application_id),
                    [
                        str(
                            application_knowledge_mapping.source_id
                        ) for application_knowledge_mapping in
                        QuerySet(ResourceMapping).filter(source_type='APPLICATION',
                                                         target_type='KNOWLEDGE',
                                                         target_id=self.data.get('knowledge_id'))
                    ]
                ))
            }

        @transaction.atomic
        def edit(self, instance: Dict, select_one=True):
            self.is_valid()
            knowledge = QuerySet(Knowledge).get(id=self.data.get("knowledge_id"))
            KnowledgeEditRequest(data=instance).is_valid(knowledge=knowledge)
            if 'embedding_model_id' in instance:
                knowledge.embedding_model_id = instance.get('embedding_model_id')
            if "name" in instance:
                knowledge.name = instance.get("name")
            if 'desc' in instance:
                knowledge.desc = instance.get("desc")
            if 'meta' in instance:
                knowledge.meta = instance.get('meta')
            if 'folder_id' in instance:
                knowledge.folder_id = instance.get('folder_id')
            if 'file_size_limit' in instance:
                knowledge.file_size_limit = instance.get('file_size_limit')
            if 'file_count_limit' in instance:
                knowledge.file_count_limit = instance.get('file_count_limit')
            knowledge.save()
            update_resource_mapping_by_knowledge(str(knowledge.id))
            if select_one:
                return self.one()
            return None

        @transaction.atomic
        def delete(self):
            self.is_valid()
            knowledge = QuerySet(Knowledge).get(id=self.data.get("knowledge_id"))
            QuerySet(Document).filter(knowledge=knowledge).delete()
            QuerySet(ProblemParagraphMapping).filter(knowledge=knowledge).delete()
            QuerySet(Paragraph).filter(knowledge=knowledge).delete()
            QuerySet(Problem).filter(knowledge=knowledge).delete()
            QuerySet(WorkspaceUserResourcePermission).filter(target=knowledge.id).delete()
            drop_knowledge_index(knowledge_id=knowledge.id)
            knowledge.delete()
            File.objects.filter(
                source_id=knowledge.id,
            ).delete()
            QuerySet(ResourceMapping).filter(
                Q(target_id=self.data.get('knowledge_id')) | Q(source_id=self.data.get('knowledge_id'))
            ).delete()
            delete_embedding_by_knowledge(self.data.get('knowledge_id'))
            return True

        def export_excel(self, with_valid=True):
            if with_valid:
                self.is_valid(raise_exception=True)
            document_list = QuerySet(Document).filter(knowledge_id=self.data.get('knowledge_id'))
            paragraph_list = native_search(
                QuerySet(Paragraph).filter(knowledge_id=self.data.get("knowledge_id")),
                get_file_content(
                    os.path.join(PROJECT_DIR, "apps", "knowledge", 'sql', 'list_paragraph_document_name.sql')
                )
            )
            problem_mapping_list = native_search(
                QuerySet(ProblemParagraphMapping).filter(knowledge_id=self.data.get("knowledge_id")),
                get_file_content(os.path.join(PROJECT_DIR, "apps", "knowledge", 'sql', 'list_problem_mapping.sql')),
                with_table_name=True
            )
            data_dict, document_dict = DocumentSerializers.Operate.merge_problem(
                paragraph_list, problem_mapping_list, document_list
            )
            workbook = DocumentSerializers.Operate.get_workbook(data_dict, document_dict)
            response = HttpResponse(content_type='application/vnd.ms-excel')
            response['Content-Disposition'] = 'attachment; filename="knowledge.xlsx"'
            workbook.save(response)
            return response

        def export_zip(self, with_valid=True):
            if with_valid:
                self.is_valid(raise_exception=True)
            knowledge = QuerySet(Knowledge).filter(id=self.data.get("knowledge_id")).first()
            document_list = QuerySet(Document).filter(knowledge_id=self.data.get('knowledge_id'))
            paragraph_list = native_search(
                QuerySet(Paragraph).filter(knowledge_id=self.data.get("knowledge_id")),
                get_file_content(
                    os.path.join(PROJECT_DIR, "apps", "knowledge", 'sql', 'list_paragraph_document_name.sql')
                )
            )
            problem_mapping_list = native_search(
                QuerySet(ProblemParagraphMapping).filter(knowledge_id=self.data.get("knowledge_id")),
                get_file_content(os.path.join(PROJECT_DIR, "apps", "knowledge", 'sql', 'list_problem_mapping.sql')),
                with_table_name=True
            )
            data_dict, document_dict = DocumentSerializers.Operate.merge_problem(
                paragraph_list, problem_mapping_list, document_list
            )
            res = [parse_image(paragraph.get('content')) for paragraph in paragraph_list]

            workbook = DocumentSerializers.Operate.get_workbook(data_dict, document_dict)
            response = HttpResponse(content_type='application/zip')
            response['Content-Disposition'] = f'attachment; filename="{knowledge.name}.zip"'
            zip_buffer = io.BytesIO()
            with TemporaryDirectory() as tempdir:
                knowledge_file = os.path.join(tempdir, 'knowledge.xlsx')
                workbook.save(knowledge_file)
                for r in res:
                    write_image(tempdir, r)
                zip_dir(tempdir, zip_buffer)
            response.write(zip_buffer.getvalue())
            return response

        def export_knowledge(self, with_valid=True):
            if with_valid:
                self.is_valid(raise_exception=True)
            knowledge_id = self.data.get("knowledge_id")
            knowledge = QuerySet(Knowledge).filter(id=knowledge_id).first()

            document_list = QuerySet(Document).filter(knowledge_id=knowledge_id)
            paragraph_list = native_search(
                QuerySet(Paragraph).filter(knowledge_id=self.data.get("knowledge_id")),
                get_file_content(
                    os.path.join(PROJECT_DIR, "apps", "knowledge", 'sql', 'list_paragraph_document_name.sql')
                )
            )
            problem_mapping_list = native_search(
                QuerySet(ProblemParagraphMapping).filter(knowledge_id=self.data.get("knowledge_id")),
                get_file_content(os.path.join(PROJECT_DIR, "apps", "knowledge", 'sql', 'list_problem_mapping.sql')),
                with_table_name=True
            )
            data_dict, document_dict = DocumentSerializers.Operate.merge_problem(
                paragraph_list, problem_mapping_list, document_list
            )

            # 查询标签和文档标签关联
            tag_list = list(QuerySet(Tag).filter(knowledge_id=knowledge_id).values('id', 'key', 'value'))
            document_tag_list = list(
                QuerySet(DocumentTag).filter(document__knowledge_id=knowledge_id).values('document_id', 'tag_id')
            )
            # 知识库标签map
            tag_map = {t['id']: t for t in tag_list}
            # 文档标签map
            doc_tag_map = defaultdict(list)

            for dt in document_tag_list:
                tag = tag_map.get(dt['tag_id'])
                if tag:
                    doc_tag_map[dt['document_id']].append(f"{tag['key']}:{tag['value']}")

            # doc_id -> document_obj
            doc_obj_map = {doc.id: doc for doc in document_list}

            # paragraph_id -> is_active
            paragraph_active_map = {}
            for p in paragraph_list:
                doc_id = p.get('document_id')
                if doc_id not in paragraph_active_map:
                    paragraph_active_map[doc_id] = []
                paragraph_active_map[doc_id].append('1' if p.get('is_active') else '0')

            res = [parse_image(paragraph.get('content')) for paragraph in paragraph_list]
            # 新增字段
            workbook = self._get_knowledge_workbook(data_dict, document_dict, doc_tag_map, doc_obj_map,
                                                    paragraph_active_map)

            response = HttpResponse(content_type='application/zip')
            response['Content-Disposition'] = f"attachment; filename*=UTF-8''{quote(knowledge.name)}.zip"
            zip_buffer = io.BytesIO()
            with TemporaryDirectory() as tempdir:
                knowledge_file_path = os.path.join(tempdir, 'knowledge.xlsx')
                workbook.save(knowledge_file_path)

                for r in res:
                    write_image(tempdir, r)

                knowledge_json = {
                    'name': knowledge.name,
                    'desc': knowledge.desc,
                    'type': knowledge.type,
                    'meta': {} if knowledge.type == KnowledgeType.LARK else (knowledge.meta if knowledge.meta else {}),
                    'file_size_limit': knowledge.file_size_limit,
                    'file_count_limit': knowledge.file_count_limit,
                    'tags': [{'key': t['key'], 'value': t['value']} for t in tag_list]
                }

                with open(os.path.join(tempdir, 'knowledge.json'), 'w', encoding='utf-8') as f:
                    json.dump(knowledge_json, f, ensure_ascii=False)

                if knowledge.type == KnowledgeType.WORKFLOW:
                    knowledge_workflow = QuerySet(KnowledgeWorkflow).filter(knowledge_id=knowledge_id).first()
                    if knowledge_workflow:
                        from knowledge.serializers.knowledge_workflow import KnowledgeWorkflowSerializer
                        from knowledge.serializers.knowledge_workflow import KnowledgeWorkflowModelSerializer
                        from application.flow.tools import get_tool_id_list
                        from tools.models import Tool, ToolScope, ToolType, ToolWorkflow
                        from knowledge.serializers.knowledge_workflow import KBWFInstance

                        tool_id_list = get_tool_id_list(knowledge_workflow.work_flow, True)
                        tool_list = []
                        if len(tool_id_list) > 0:
                            tool_list = QuerySet(Tool).filter(id__in=tool_id_list).exclude(scope=ToolScope.SHARED)
                        tw_dict = {tw.tool_id: tw
                                   for tw in QuerySet(ToolWorkflow).filter(
                                tool_id__in=[tool.id for tool in tool_list if tool.tool_type == ToolType.WORKFLOW])}
                        knowledge_workflow_dict = KnowledgeWorkflowModelSerializer(knowledge_workflow).data

                        kbwf_instance = KBWFInstance(
                            knowledge_workflow_dict,
                            [],
                            'v2',
                            [KnowledgeWorkflowSerializer.Export.to_tool_dict(tool, tw_dict) for tool in tool_list]
                        )
                        knowledge_workflow_pickle = pickle.dumps(kbwf_instance)
                        with open(os.path.join(tempdir, 'workflow.kbwf'), 'wb') as f:
                            f.write(knowledge_workflow_pickle)
                zip_dir(tempdir, zip_buffer)
            response.write(zip_buffer.getvalue())
            return response

        @staticmethod
        def _get_knowledge_workbook(data_dict: dict, document_dict: dict, doc_tag_map: dict, doc_obj_map: dict,
                                    paragraph_active_map: dict):
            import openpyxl
            from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
            workbook = openpyxl.Workbook()
            workbook.remove(workbook.active)
            if len(data_dict.keys()) == 0:
                data_dict['sheet'] = []
            for sheet_id in data_dict:
                sheet_name = document_dict.get(sheet_id)
                worksheet = workbook.create_sheet(sheet_name)

                doc = doc_obj_map.get(sheet_id) if sheet_id in doc_obj_map else None
                tags_str = '|'.join(doc_tag_map.get(sheet_id, []))
                hit_method = doc.hit_handling_method if doc else ''
                similarity = doc.directly_return_similarity if doc else ''
                is_active = '1' if (doc and doc.is_active) else '0'
                doc_type = doc.type if doc else ''
                doc_meta = json.dumps(doc.meta, ensure_ascii=False) if (doc and doc.meta) else ''

                header = [gettext('Section title (optional)'),
                          gettext('Section content (required, question answer, no more than 4096 characters)'),
                          gettext('Question (optional, one per line in the cell)'),
                          gettext('Tags'),
                          gettext('Hit handling method'),
                          gettext('Directly return similarity'),
                          gettext('Is active'),
                          gettext('Paragraph is active'),
                          gettext('Document type'),
                          gettext('Document meta')]

                rows = data_dict.get(sheet_id, [])
                para_active_list = paragraph_active_map.get(sheet_id, [])
                # 初始化标题
                data = [header]
                for row_idx, row in enumerate(rows):
                    para_active = para_active_list[row_idx] if row_idx < len(para_active_list) else '1'
                    # None 转为 ''
                    row = [col if col is not None else '' for col in row]
                    # 补齐到3列
                    row = (row + ['','',''])[:3]
                    if row_idx == 0:
                        data.append(
                            [*row, tags_str, hit_method, similarity, is_active, para_active, doc_type, doc_meta])
                    else:
                        data.append([*row, '', '', '', '', para_active, '', ''])

                for row_idx, row in enumerate(data):
                    for col_idx, col in enumerate(row):
                        cell = worksheet.cell(row=row_idx + 1, column=col_idx + 1)
                        if isinstance(col, str):
                            col = re.sub(ILLEGAL_CHARACTERS_RE, '', col)
                            if col.startswith(('=', '+', '-', '@')):
                                col = '\ufeff' + col
                        cell.value = col
            return workbook

        @staticmethod
        def merge_problem(paragraph_list: List[Dict], problem_mapping_list: List[Dict]):
            result = {}
            document_dict = {}

            for paragraph in paragraph_list:
                problem_list = [problem_mapping.get('content') for problem_mapping in problem_mapping_list if
                                problem_mapping.get('paragraph_id') == paragraph.get('id')]
                document_sheet = result.get(paragraph.get('document_id'))
                d = document_dict.get(paragraph.get('document_name'))
                if d is None:
                    document_dict[paragraph.get('document_name')] = {paragraph.get('document_id')}
                else:
                    d.add(paragraph.get('document_id'))

                if document_sheet is None:
                    result[paragraph.get('document_id')] = [[paragraph.get('title'), paragraph.get('content'),
                                                             '\n'.join(problem_list)]]
                else:
                    document_sheet.append([paragraph.get('title'), paragraph.get('content'), '\n'.join(problem_list)])
            result_document_dict = {}
            for d_name in document_dict:
                for index, d_id in enumerate(document_dict.get(d_name)):
                    result_document_dict[d_id] = d_name if index == 0 else d_name + str(index)
            return result, result_document_dict

    class ImportKnowledge(serializers.Serializer):
        user_id = serializers.UUIDField(required=True, label=_('user id'))
        workspace_id = serializers.CharField(required=True, label=_('workspace id'))
        folder_id = serializers.CharField(required=True, label=_('folder id'))

        @transaction.atomic
        def import_knowledge(self, file, is_import_tool=False, with_valid=True):
            if with_valid:
                self.is_valid(raise_exception=True)
                KnowledgeImportRequest(data={'file': file}).is_valid(raise_exception=True)

            try:
                zf = zipfile.ZipFile(file)
            except zipfile.BadZipFile:
                raise AppApiException(500, _('Not a valid zip file'))

            namelist = zf.namelist()
            if 'knowledge.json' not in namelist:
                raise AppApiException(500, _('Not a valid KB export file, missing knowledge.json'))
            if 'knowledge.xlsx' not in namelist:
                raise AppApiException(500, _('Not a valid KB export file, missing knowledge.xlsx'))

            # knowledge.json -> knowledge
            knowledge_data = json.loads(zf.read('knowledge.json'))
            workspace_id = self.data.get('workspace_id')
            user_id = self.data.get('user_id')
            knowledge_id = uuid.uuid7()
            folder_id = self.data.get('folder_id')
            knowledge = Knowledge(
                id=knowledge_id,
                name=knowledge_data.get('name', 'Untitled'),
                desc=knowledge_data.get('desc', ''),
                type=knowledge_data.get('type', KnowledgeType.BASE),
                scope=self.data.get('scope', KnowledgeScope.WORKSPACE),
                meta=knowledge_data.get('meta', {}),
                file_size_limit=knowledge_data.get('file_size_limit', 100),
                file_count_limit=knowledge_data.get('file_count_limit', 50),
                embedding_model=None,
                user_id=user_id,
                workspace_id=workspace_id,
                folder_id=folder_id
            )
            knowledge.save()

            # 图片
            old_to_new_file_map = {}
            for name in namelist:
                if name.startswith('oss/file/') and name != 'oss/file/':
                    old_id = name.split('/')[-1]
                    if not old_id:
                        continue
                    file_bytes = zf.read(name)
                    new_file = File(
                        id=uuid.uuid7(),
                        file_name=old_id,
                        source_type=FileSourceType.KNOWLEDGE,
                        source_id=str(knowledge_id),
                        meta={}
                    )
                    new_file.save(bytea=file_bytes)
                    old_to_new_file_map[old_id] = str(new_file.id)

            # knowledge.xlsx -> doc + para + problem
            import openpyxl
            xlsx_bytes = io.BytesIO(zf.read('knowledge.xlsx'))
            workbook = openpyxl.load_workbook(xlsx_bytes)

            document_model_list = []
            paragraph_model_list = []
            problem_paragraph_object_list = []
            doc_tags_map = {}

            for sheet in workbook.worksheets:
                doc_name = sheet.title
                rows = list(sheet.iter_rows(min_row=2, values_only=True))
                if not rows:
                    continue

                # 首行文档元数据
                first_row = rows[0]
                tags_str = first_row[3] if len(first_row) > 3 and first_row[3] else ''
                hit_method = first_row[4] if len(first_row) > 4 and first_row[4] else 'optimization'
                similarity = first_row[5] if len(first_row) > 5 and first_row[5] else 0.9
                doc_is_active = first_row[6] if len(first_row) > 6 and first_row[6] else '1'
                doc_type = first_row[8] if len(first_row) > 8 and first_row[8] else knowledge_data.get('type',
                                                                                                       KnowledgeType.BASE)
                doc_meta_str = first_row[9] if len(first_row) > 9 and first_row[9] else '{}'

                try:
                    doc_meta = json.loads(doc_meta_str) if isinstance(doc_meta_str, str) else {}
                except (json.JSONDecodeError, TypeError):
                    doc_meta = {}

                char_length = sum(len(row[1] or '') for row in rows)
                document_id = uuid.uuid7()
                document = Document(
                    id=document_id,
                    knowledge_id=knowledge_id,
                    name=doc_name,
                    char_length=char_length,
                    is_active=str(doc_is_active) == '1',
                    type=doc_type,
                    hit_handling_method=hit_method,
                    directly_return_similarity=float(similarity) if similarity else 0.9,
                    meta=doc_meta
                )

                document_model_list.append(document)
                if tags_str:
                    doc_tags_map[document_id] = tags_str
                # 逐行创建 para + problem
                for row_idx, row in enumerate(rows):
                    title = str(row[0]) if len(row) > 0 and row[0] is not None else ''
                    content = str(row[1]) if len(row) > 1 and row[1] is not None else ''
                    problems_str = str(row[2]) if len(row) > 2 and row[2] is not None else ''
                    para_is_active = row[7] if len(row) > 7 and row[7] else '1'

                    # 图片 link 替换
                    for old_id, new_id in old_to_new_file_map.items():
                        content = content.replace(old_id, new_id)

                    if title.startswith('\ufeff'):
                        title = title[1:]
                    if content.startswith('\ufeff'):
                        content = content[1:]

                    paragraph_id = uuid.uuid7()
                    paragraph = Paragraph(
                        id=paragraph_id,
                        document_id=document_id,
                        knowledge_id=knowledge_id,
                        title=title,
                        content=content,
                        is_active=str(para_is_active) == '1',
                        position=row_idx + 1
                    )
                    paragraph_model_list.append(paragraph)

                    if problems_str:
                        if problems_str.startswith('\ufeff'):
                            problems_str = problems_str[1:]
                        for problem_content in problems_str.split('\n'):
                            problem_content = problem_content.strip()
                            if problem_content:
                                problem_paragraph_object_list.append(ProblemParagraphObject(
                                    knowledge_id, document_id, paragraph_id, problem_content
                                ))
            # bulk create
            QuerySet(Document).bulk_create(document_model_list) if len(document_model_list) > 0 else None
            QuerySet(Paragraph).bulk_create(paragraph_model_list) if len(paragraph_model_list) > 0 else None

            # 问题
            problem_model_list, problem_paragraph_mapping_list = (
                ProblemParagraphManage(problem_paragraph_object_list, knowledge_id).to_problem_model_list()
            )
            bulk_create_in_batches(Problem, problem_model_list, batch_size=1000)
            bulk_create_in_batches(ProblemParagraphMapping, problem_paragraph_mapping_list, batch_size=1000)

            # Tag
            tag_list = knowledge_data.get('tags', [])
            if tag_list:
                tag_model_list = []
                tag_key_value_to_model = {}
                for tag in tag_list:
                    tag_model = Tag(
                        id=uuid.uuid7(),
                        knowledge_id=knowledge_id,
                        key=tag['key'],
                        value=tag['value']
                    )
                    tag_model_list.append(tag_model)

                    tag_key_value_to_model[f"{tag['key']}:{tag['value']}"] = tag_model
                QuerySet(Tag).bulk_create(tag_model_list)

                # Document_Tag
                document_tag_model_list = []
                for doc_id, tags_str in doc_tags_map.items():
                    for tag_str in tags_str.split('|'):
                        tag_str = tag_str.strip()
                        if tag_str and tag_str in tag_key_value_to_model:
                            document_tag_model_list.append(DocumentTag(
                                id=uuid.uuid7(),
                                document_id=doc_id,
                                tag_id=tag_key_value_to_model[tag_str].id
                            ))
                QuerySet(DocumentTag).bulk_create(document_tag_model_list) if len(document_tag_model_list) > 0 else None

                # 工作流导入
            if 'workflow.kbwf' in namelist:
                workflow_bytes = zf.read('workflow.kbwf')
                from knowledge.serializers.knowledge_workflow import KnowledgeWorkflowSerializer
                workflow_file = SimpleUploadedFile('workflow.kbwf', workflow_bytes)
                KnowledgeWorkflowSerializer.Import(
                    data={'knowledge_id': str(knowledge_id), 'user_id': user_id, 'workspace_id': workspace_id}
                ).import_({'file': workflow_file}, is_import_tool)

            # 授权 + 资源映射
            UserResourcePermissionSerializer(data={
                'workspace_id': self.data.get('workspace_id'),
                'user_id': self.data.get('user_id'),
                'auth_target_type': AuthTargetType.KNOWLEDGE.value
            }).auth_resource(str(knowledge_id))

            update_resource_mapping_by_knowledge(str(knowledge_id))

            zf.close()
            return {'knowledge_id': str(knowledge_id), 'type': knowledge.type}

    class Create(serializers.Serializer):
        user_id = serializers.UUIDField(required=True, label=_('user id'))
        workspace_id = serializers.CharField(required=True, label=_('workspace id'))
        scope = serializers.ChoiceField(required=False, label=_('scope'), default=KnowledgeScope.WORKSPACE,
                                        choices=KnowledgeScope.choices)

        @staticmethod
        def post_embedding_knowledge(document_list, knowledge_id):
            model_id = get_embedding_model_id_by_knowledge_id(knowledge_id)
            embedding_by_knowledge.delay(knowledge_id, model_id)
            return document_list

        @post(post_function=post_embedding_knowledge)
        @transaction.atomic
        def save_base(self, instance, with_valid=True):
            if with_valid:
                self.is_valid(raise_exception=True)
                KnowledgeBaseCreateRequest(data=instance).is_valid(raise_exception=True)
            folder_id = instance.get('folder_id', self.data.get('workspace_id'))

            knowledge_id = uuid.uuid7()
            knowledge = Knowledge(
                id=knowledge_id,
                name=instance.get('name'),
                workspace_id=self.data.get('workspace_id'),
                desc=instance.get('desc'),
                type=instance.get('type', KnowledgeType.BASE),
                user_id=self.data.get('user_id'),
                scope=self.data.get('scope', KnowledgeScope.WORKSPACE),
                folder_id=folder_id,
                embedding_model_id=instance.get('embedding_model_id'),
                meta=instance.get('meta', {}),
            )

            document_model_list = []
            paragraph_model_list = []
            problem_paragraph_object_list = []
            # 插入文档
            for document in instance.get('documents') if 'documents' in instance else []:
                document_paragraph_dict_model = DocumentSerializers.Create.get_document_paragraph_model(knowledge_id,
                                                                                                        document)
                document_model_list.append(document_paragraph_dict_model.get('document'))
                for paragraph in document_paragraph_dict_model.get('paragraph_model_list'):
                    paragraph_model_list.append(paragraph)
                for problem_paragraph_object in document_paragraph_dict_model.get('problem_paragraph_object_list'):
                    problem_paragraph_object_list.append(problem_paragraph_object)

            problem_model_list, problem_paragraph_mapping_list = (
                ProblemParagraphManage(problem_paragraph_object_list, knowledge_id)
                .to_problem_model_list())
            # 插入知识库
            knowledge.save()
            # 插入文档
            QuerySet(Document).bulk_create(document_model_list) if len(document_model_list) > 0 else None
            # 批量插入段落
            QuerySet(Paragraph).bulk_create(paragraph_model_list) if len(paragraph_model_list) > 0 else None
            # 批量插入问题
            QuerySet(Problem).bulk_create(problem_model_list) if len(problem_model_list) > 0 else None
            # 批量插入关联问题
            QuerySet(ProblemParagraphMapping).bulk_create(
                problem_paragraph_mapping_list
            ) if len(problem_paragraph_mapping_list) > 0 else None
            # 自动资源给授权当前用户
            UserResourcePermissionSerializer(data={
                'workspace_id': self.data.get('workspace_id'),
                'user_id': self.data.get('user_id'),
                'auth_target_type': AuthTargetType.KNOWLEDGE.value
            }).auth_resource(str(knowledge_id))
            update_resource_mapping_by_knowledge(str(knowledge_id))
            return {
                **KnowledgeModelSerializer(knowledge).data,
                'user_id': self.data.get('user_id'),
                'document_list': document_model_list,
                "document_count": len(document_model_list),
                "char_length": reduce(lambda x, y: x + y, [d.char_length for d in document_model_list], 0)
            }, knowledge_id

        def save_web(self, instance: Dict, with_valid=True):
            if with_valid:
                self.is_valid(raise_exception=True)
                KnowledgeWebCreateRequest(data=instance).is_valid(raise_exception=True)

            folder_id = instance.get('folder_id', self.data.get('workspace_id'))

            knowledge_id = uuid.uuid7()
            knowledge = Knowledge(
                id=knowledge_id,
                name=instance.get('name'),
                desc=instance.get('desc'),
                user_id=self.data.get('user_id'),
                type=instance.get('type', KnowledgeType.WEB),
                scope=self.data.get('scope', KnowledgeScope.WORKSPACE),
                folder_id=folder_id,
                workspace_id=self.data.get('workspace_id'),
                embedding_model_id=instance.get('embedding_model_id'),
                meta={
                    'source_url': instance.get('source_url'),
                    'selector': instance.get('selector', 'body'),
                    'embedding_model_id': instance.get('embedding_model_id')
                },
            )
            knowledge.save()
            # 自动资源给授权当前用户
            UserResourcePermissionSerializer(data={
                'workspace_id': self.data.get('workspace_id'),
                'user_id': self.data.get('user_id'),
                'auth_target_type': AuthTargetType.KNOWLEDGE.value
            }).auth_resource(str(knowledge_id))

            sync_web_knowledge.delay(str(knowledge_id), instance.get('source_url'), instance.get('selector'))
            update_resource_mapping_by_knowledge(str(knowledge_id))
            return {**KnowledgeModelSerializer(knowledge).data, 'document_list': []}

    class SyncWeb(serializers.Serializer):
        workspace_id = serializers.CharField(required=True, label=_('workspace id'))
        knowledge_id = serializers.CharField(required=True, label=_('knowledge id'))
        user_id = serializers.UUIDField(required=False, label=_('user id'))
        sync_type = serializers.CharField(required=True, label=_('sync type'), validators=[
            validators.RegexValidator(regex=re.compile("^replace|complete$"),
                                      message=_('The synchronization type only supports:replace|complete'), code=500)])

        def is_valid(self, *, raise_exception=False):
            super().is_valid(raise_exception=True)
            workspace_id = self.data.get('workspace_id')
            query_set = QuerySet(Knowledge).filter(id=self.data.get('knowledge_id'))
            if workspace_id:
                query_set = query_set.filter(workspace_id=workspace_id)
            if not query_set.exists():
                raise AppApiException(500, _('Knowledge id does not exist'))
            first = QuerySet(Knowledge).filter(id=self.data.get("knowledge_id")).first()
            if first is None:
                raise AppApiException(300, _('id does not exist'))
            if first.type != KnowledgeType.WEB:
                raise AppApiException(500, _('Synchronization is only supported for web site types'))

        def sync(self, with_valid=True):
            if with_valid:
                self.is_valid(raise_exception=True)
            sync_type = self.data.get('sync_type')
            knowledge_id = self.data.get('knowledge_id')
            knowledge = QuerySet(Knowledge).get(id=knowledge_id)
            self.__getattribute__(sync_type + '_sync')(knowledge)
            return True

        @staticmethod
        def get_sync_handler(knowledge):
            def handler(child_link: ChildLink, response: Fork.Response):
                if response.status == 200:
                    try:
                        document_name = child_link.tag.text if child_link.tag is not None and len(
                            child_link.tag.text.strip()) > 0 else child_link.url
                        paragraphs = get_split_model('web.md').parse(response.content)
                        maxkb_logger.info(child_link.url.strip())
                        first = QuerySet(Document).filter(
                            meta__source_url=child_link.url.strip(),
                            knowledge=knowledge
                        ).first()
                        if first is not None:
                            # 如果存在,使用文档同步
                            DocumentSerializers.Sync(data={'document_id': first.id}).sync()
                        else:
                            # 插入
                            DocumentSerializers.Create(data={'knowledge_id': knowledge.id}).save(
                                {'name': document_name, 'paragraphs': paragraphs,
                                 'meta': {'source_url': child_link.url.strip(),
                                          'selector': knowledge.meta.get('selector')},
                                 'type': KnowledgeType.WEB}, with_valid=True)
                    except Exception as e:
                        maxkb_logger.error(f'{str(e)}:{traceback.format_exc()}')

            return handler

        def replace_sync(self, knowledge):
            """
            替换同步
            :return:
            """
            url = knowledge.meta.get('source_url')
            selector = knowledge.meta.get('selector') if 'selector' in knowledge.meta else None
            sync_replace_web_knowledge.delay(str(knowledge.id), url, selector)

        def complete_sync(self, knowledge):
            """
            完整同步  删掉当前数据集下所有的文档,再进行同步
            :return:
            """
            # 删除关联问题
            QuerySet(ProblemParagraphMapping).filter(knowledge=knowledge).delete()
            # 删除文档
            QuerySet(Document).filter(knowledge=knowledge).delete()
            # 删除段落
            QuerySet(Paragraph).filter(knowledge=knowledge).delete()
            # 删除向量
            delete_embedding_by_knowledge(self.data.get('knowledge_id'))
            # 同步
            self.replace_sync(knowledge)

    class HitTest(serializers.Serializer):
        workspace_id = serializers.CharField(required=True, label=_('workspace id'))
        knowledge_id = serializers.UUIDField(required=True, label=_("id"))
        user_id = serializers.UUIDField(required=False, label=_('user id'))
        query_text = serializers.CharField(required=True, label=_('query text'))
        top_number = serializers.IntegerField(required=True, max_value=10000, min_value=1, label=_("top number"))
        similarity = serializers.FloatField(required=True, max_value=2, min_value=0, label=_('similarity'))
        search_mode = serializers.CharField(required=True, label=_('search mode'), validators=[
            validators.RegexValidator(regex=re.compile("^embedding|keywords|blend$"),
                                      message=_('The type only supports embedding|keywords|blend'), code=500)
        ])

        def is_valid(self, *, raise_exception=True):
            super().is_valid(raise_exception=True)
            workspace_id = self.data.get('workspace_id')
            query_set = QuerySet(Knowledge).filter(id=self.data.get('knowledge_id'))
            if workspace_id:
                query_set = query_set.filter(workspace_id=workspace_id)
            if not query_set.exists():
                raise AppApiException(500, _('Knowledge id does not exist'))
            if not QuerySet(Knowledge).filter(id=self.data.get("knowledge_id")).exists():
                raise AppApiException(300, _('id does not exist'))

        def hit_test(self):
            self.is_valid()
            vector = VectorStore.get_embedding_vector()
            exclude_document_id_list = [
                str(
                    document.id
                ) for document in QuerySet(Document).filter(knowledge_id=self.data.get('knowledge_id'), is_active=False)
            ]
            model = get_embedding_model_by_knowledge_id(self.data.get('knowledge_id'))
            # 向量库检索
            hit_list = vector.hit_test(
                self.data.get('query_text'),
                [self.data.get('knowledge_id')],
                exclude_document_id_list,
                self.data.get('top_number'),
                self.data.get('similarity'),
                SearchMode(self.data.get('search_mode')),
                model
            )
            hit_dict = reduce(lambda x, y: {**x, **y}, [{hit.get('paragraph_id'): hit} for hit in hit_list], {})
            p_list = list_paragraph([h.get('paragraph_id') for h in hit_list])
            return [
                {
                    **p,
                    'similarity': hit_dict.get(p.get('id')).get('similarity'),
                    'comprehensive_score': hit_dict.get(p.get('id')).get('comprehensive_score')
                } for p in p_list
            ]

    class StoreKnowledge(serializers.Serializer):
        user_id = serializers.UUIDField(required=True, label=_("User ID"))
        name = serializers.CharField(required=False, label=_("tool name"), allow_null=True, allow_blank=True)

        def get_appstore_templates(self):
            self.is_valid(raise_exception=True)
            # 下载zip文件
            try:
                appstore_url = CONFIG.get('APPSTORE_URL', 'https://apps-assets.fit2cloud.com/stable/maxkb.json.zip')
                res = requests.get(appstore_url, timeout=5)
                res.raise_for_status()
                # 创建临时文件保存zip
                with tempfile.NamedTemporaryFile(delete=False, suffix='.zip') as temp_zip:
                    temp_zip.write(res.content)
                    temp_zip_path = temp_zip.name

                try:
                    # 解压zip文件
                    with zipfile.ZipFile(temp_zip_path, 'r') as zip_ref:
                        # 获取zip中的第一个文件（假设只有一个json文件）
                        json_filename = zip_ref.namelist()[0]
                        json_content = zip_ref.read(json_filename)

                    # 将json转换为字典
                    tool_store = json.loads(json_content.decode('utf-8'))
                    tag_dict = {tag['name']: tag['key'] for tag in tool_store['additionalProperties']['tags']}
                    filter_apps = []
                    for tool in tool_store['apps']:
                        if self.data.get('name', '') != '':
                            if self.data.get('name').lower() not in tool.get('name', '').lower():
                                continue
                        if not tool['downloadUrl'].endswith('.kbwf'):
                            continue
                        versions = tool.get('versions', [])
                        tool['label'] = tag_dict[tool.get('tags')[0]] if tool.get('tags') else ''
                        tool['version'] = next(
                            (version.get('name') for version in versions if
                             version.get('downloadUrl') == tool['downloadUrl']),
                        )
                        filter_apps.append(tool)

                    tool_store['apps'] = filter_apps
                    return tool_store
                finally:
                    # 清理临时文件
                    os.unlink(temp_zip_path)
            except Exception as e:
                maxkb_logger.error(f"fetch appstore tools error: {e}")
                return {'apps': [], 'additionalProperties': {'tags': []}}

    class TransformWorkflow(serializers.Serializer):
        workspace_id = serializers.CharField(required=True, label=_('Workspace ID'))
        knowledge_id = serializers.UUIDField(required=True, label=_('Knowledge ID'))
        user_id = serializers.UUIDField(required=True, label=_('User ID'))

        def transform(self, instance: Dict):
            self.is_valid(raise_exception=True)
            knowledge = QuerySet(Knowledge).filter(
                id=self.data.get('knowledge_id'),
                workspace_id=self.data.get('workspace_id')
            ).first()

            if not knowledge:
                raise AppApiException(500, _('Knowledge not found'))
            if knowledge.type == KnowledgeType.WORKFLOW:
                raise AppApiException(500, _('Knowledge is already a workflow'))

            knowledge.type = KnowledgeType.WORKFLOW
            knowledge.save()

            workflow_id = uuid.uuid7()
            knowledge_workflow = KnowledgeWorkflow(
                id=workflow_id,
                workspace_id=knowledge.workspace_id,
                knowledge_id=knowledge.id,
                work_flow=instance.get('work_flow', {}),
            )
            knowledge_workflow.save()
            return True

    class Tags(serializers.Serializer):
        workspace_id = serializers.CharField(required=True, label=_('workspace id'))
        user_id = serializers.UUIDField(required=True, label=_('user id'))
        knowledge_ids = serializers.ListField(
            required=True, label=_('knowledge ids'),
            child=serializers.UUIDField(required=True, label=_('id'))
        )

        def list(self):
            self.is_valid(raise_exception=True)
            if self.data.get('name'):
                name = self.data.get('name')
                tags = QuerySet(Tag).filter(
                    knowledge_id__in=self.data.get('knowledge_ids')
                ).filter(
                    Q(key__icontains=name) | Q(value__icontains=name)
                ).values('key', 'value', 'id', 'create_time', 'update_time').order_by('create_time', 'key', 'value')
            else:
                # 获取所有标签，按创建时间排序保持稳定顺序
                tags = QuerySet(Tag).filter(
                    knowledge_id__in=self.data.get('knowledge_ids')
                ).values('key', 'value', 'id', 'create_time', 'update_time').order_by('create_time', 'key', 'value')

            # 按key分组
            grouped_tags = defaultdict(list)
            for tag in tags:
                grouped_tags[tag['key']].append({
                    'id': tag['id'],
                    'value': tag['value'],
                    'create_time': tag['create_time'],
                    'update_time': tag['update_time']
                })

            # 转换为期望的格式，保持key的顺序
            result = []
            # 按key排序以确保结果顺序一致
            for key in sorted(grouped_tags.keys()):
                values = grouped_tags[key]
                # 按创建时间对values进行排序
                values.sort(key=lambda x: x['create_time'])
                result.append({
                    'key': key,
                    'values': values,
                })

            return result


class KnowledgeBatchOperateSerializer(serializers.Serializer):
    workspace_id = serializers.CharField(required=True, label=_('workspace id'))

    def is_valid(self, *, raise_exception=False):
        super().is_valid(raise_exception=True)

    @transaction.atomic
    def batch_delete(self, instance: Dict, with_valid=True):
        if with_valid:
            BatchSerializer(data=instance).is_valid(model=Knowledge, raise_exception=True)
            self.is_valid(raise_exception=True)
        id_list = instance.get('id_list')
        workspace_id = self.data.get('workspace_id')
        knowledge_query_set = QuerySet(Knowledge).filter(id__in=id_list, workspace_id=workspace_id)

        # 删除所有关联
        QuerySet(Document).filter(knowledge__in=knowledge_query_set).delete()
        QuerySet(ProblemParagraphMapping).filter(knowledge__in=knowledge_query_set).delete()
        QuerySet(Paragraph).filter(knowledge__in=knowledge_query_set).delete()
        QuerySet(Problem).filter(knowledge__in=knowledge_query_set).delete()
        QuerySet(WorkspaceUserResourcePermission).filter(target__in=id_list).delete()

        for k_id in id_list:
            drop_knowledge_index(knowledge_id=k_id)
            delete_embedding_by_knowledge(k_id)

        File.objects.filter(source_id__in=id_list).delete()
        QuerySet(ResourceMapping).filter(
            Q(target_id__in=id_list) | Q(source_id__in=id_list)
        ).delete()

        knowledge_query_set.delete()
        return True

    def batch_move(self, instance: Dict, with_valid=True):
        if with_valid:
            BatchMoveSerializer(data=instance).is_valid(model=Knowledge, raise_exception=True)
            self.is_valid(raise_exception=True)
        id_list = instance.get('id_list')
        folder_id = instance.get('folder_id')
        workspace_id = self.data.get('workspace_id')

        QuerySet(Knowledge).filter(id__in=id_list, workspace_id=workspace_id).update(folder_id=folder_id)
        return True


